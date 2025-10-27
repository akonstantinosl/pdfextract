import os
os.environ['KIVY_ASYNC_LIB'] = 'trio'
from kivymd.app import MDApp
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import MDFillRoundFlatButton, MDFillRoundFlatIconButton
from kivymd.uix.label import MDLabel
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.list import OneLineListItem
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.gridlayout import MDGridLayout
from kivymd.uix.dialog import MDDialog
from kivy.clock import mainthread, Clock
from kivy.lang import Builder
from kivy.core.window import Window
from kivy.metrics import dp
from kivy.uix.widget import Widget
from kivy.properties import ObjectProperty, StringProperty, NumericProperty
from kivy.utils import get_color_from_hex
import threading
import io
import sys
import gc
import zipfile
import subprocess
from collections import defaultdict
from datetime import datetime
from plyer import filechooser
from rapidocr_onnxruntime import RapidOCR
import numpy as np
from PIL import Image
import pdf2image
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import cv2
from scipy.signal import find_peaks
from scipy.ndimage import gaussian_filter1d

# --- Konfigurasi Path dan Direktori ---
if getattr(sys, 'frozen', False):
    # Jika dijalankan sebagai aplikasi yang di-bundle
    application_path = os.path.dirname(sys.executable)
else:
    # Jika dijalankan sebagai skrip python biasa
    application_path = os.path.dirname(os.path.abspath(__file__))

# Tambahkan path Poppler (yang di-bundle) ke environment PATH
poppler_path = os.path.join(application_path, 'poppler', 'Library', 'bin')
os.environ['PATH'] = poppler_path + os.pathsep + os.environ.get('PATH', '')

# Tentukan path absolut ke model ONNX
det_model_path = os.path.join(application_path, 'models', 'ch_PP-OCRv4_det_infer.onnx')
cls_model_path = os.path.join(application_path, 'models', 'ch_ppocr_mobile_v2.0_cls_infer.onnx')
rec_model_path = os.path.join(application_path, 'models', 'ch_PP-OCRv4_rec_infer.onnx')

# --- Patch Windows untuk Sembunyikan Console Subprocess ---
if sys.platform == 'win32':
    _original_popen = subprocess.Popen

    def _silent_popen(*args, **kwargs):
        """Wrapper untuk subprocess.Popen yang menyembunyikan console window"""
        if 'startupinfo' not in kwargs:
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = subprocess.SW_HIDE
            kwargs['startupinfo'] = startupinfo

        # Tambahkan CREATE_NO_WINDOW flag
        if 'creationflags' not in kwargs:
            kwargs['creationflags'] = 0
        kwargs['creationflags'] |= subprocess.CREATE_NO_WINDOW

        return _original_popen(*args, **kwargs)

    # Ganti subprocess.Popen dengan versi silent
    subprocess.Popen = _silent_popen

# --- Inisialisasi Engine OCR ---
print("Memuat model OCR...")
try:
    engine = RapidOCR(
        det_model_path=det_model_path,
        cls_model_path=cls_model_path,
        rec_model_path=rec_model_path,
        rec_batch_num=6,
        use_angle_cls=True,
        det_limit_side_len=960,
        det_db_thresh=0.3,
        det_db_box_thresh=0.5,
        det_db_unclip_ratio=1.6
    )
    print("Model OCR berhasil dimuat.")
except Exception as e:
    print(f"ERROR: Gagal memuat model OCR: {e}")
    # Pertimbangkan untuk menampilkan error ini di UI atau keluar dari aplikasi
    sys.exit(1)


# --- KivyMD UI Definition (KV Lang) ---
Window.clearcolor = (1, 1, 1, 1) # Atur warna latar belakang window

KV = """
#:import get_color_from_hex kivy.utils.get_color_from_hex

# --- Definisi Cell (MainScreen) ---
<GridCellLabel@MDLabel>:
    theme_text_color: "Primary"
    halign: 'left'
    valign: 'middle'
    padding_x: dp(5)
    shorten: True
    text_size: (self.width - dp(10), None) if self.width > dp(20) else (dp(10), None)
    size_hint_y: None
    height: dp(40)
    canvas.before:
        Color:
            rgba: 1, 1, 1, 1
        Rectangle:
            pos: self.pos
            size: self.size

<GridCellLabelRight@GridCellLabel>:
    halign: 'center'
    padding_x: 0
    shorten: False
    text_size: (self.width, None)

<WaitingCellWithDelete>:
    orientation: 'horizontal'
    padding: dp(5)
    spacing: dp(5)
    md_bg_color: app.theme_cls.bg_normal

    MDLabel:
        text: 'Menunggu'
        theme_text_color: "Hint"
        halign: 'center'
        valign: 'middle'
        size_hint_x: 0.7

    MDIconButton:
        icon: 'close'
        theme_text_color: "Error"
        pos_hint: {'center_y': 0.5}
        size_hint: (None, None)
        size: (dp(30), dp(30))
        on_release: app.root.main_screen.remove_file_from_list(root.file_index)

<DownloadButtonCell>:
    orientation: 'vertical'
    padding: dp(5)
    md_bg_color: app.theme_cls.bg_normal

    MDBoxLayout:
        orientation: 'horizontal'
        Widget: # Spacer Kiri
        MDFillRoundFlatIconButton:
            text: 'Unduh'
            icon: 'download'
            md_bg_color: app.theme_cls.accent_color
            size_hint: (None, None)
            size: (dp(100), dp(36))
            pos_hint: {'center_y': 0.5}
            on_release: app.root.main_screen.download_single_file(root.file_index)
        Widget: # Spacer Kanan

<FailedStatusCell@MDBoxLayout>:
    padding: dp(5)
    md_bg_color: app.theme_cls.bg_normal

    MDBoxLayout:
        orientation: 'horizontal'
        Widget: # Spacer Kiri
        MDLabel:
            text: 'Gagal'
            theme_text_color: "Error"
            bold: True
            halign: 'center'
            valign: 'middle'
            size_hint: (None, None)
            size: (dp(80), dp(36))
        Widget: # Spacer Kanan

<NoTablesStatusCell@MDBoxLayout>:
    padding: dp(5)
    md_bg_color: app.theme_cls.bg_normal

    MDBoxLayout:
        orientation: 'horizontal'
        Widget: # Spacer Kiri
        MDLabel:
            text: 'Tak Ada Tabel'
            theme_text_color: "Hint"
            halign: 'center'
            valign: 'middle'
            size_hint: (None, None)
            size: (dp(100), dp(36))
        Widget: # Spacer Kanan

<ProgressTextCell@GridCellLabelRight>:
    # Kelas ini hanya mewarisi, tidak perlu properti tambahan di sini

# --- Definisi Cell (ResultScreen) ---
<ResultCellLabel@MDLabel>:
    theme_text_color: "Primary"
    halign: 'left'
    valign: 'middle'
    padding_x: dp(5)
    shorten: True
    text_size: (self.width - dp(10), None) if self.width > dp(20) else (dp(10), None)
    size_hint_y: None
    height: dp(40)
    canvas.before:
        Color:
            rgba: 1, 1, 1, 1
        Rectangle:
            pos: self.pos
            size: self.size

<ResultCellLabelRight@ResultCellLabel>:
    halign: 'center'
    padding_x: 0
    shorten: False
    text_size: (self.width, None)

<ResultSuccessCell>:
    orientation: 'vertical'
    padding: dp(5)
    md_bg_color: app.theme_cls.bg_normal

    MDBoxLayout:
        orientation: 'horizontal'
        Widget: # Spacer Kiri
        MDFillRoundFlatIconButton:
            text: 'Unduh'
            icon: 'download'
            md_bg_color: app.theme_cls.accent_color
            size_hint: (None, None)
            size: (dp(100), dp(36))
            pos_hint: {'center_y': 0.5}
            on_release: app.root.main_screen.download_single_file(root.file_index)
        Widget: # Spacer Kanan

<ResultFailedCell@MDBoxLayout>:
    padding: dp(5)
    md_bg_color: app.theme_cls.bg_normal

    MDBoxLayout:
        orientation: 'horizontal'
        Widget: # Spacer Kiri
        MDLabel:
            text: 'Gagal'
            theme_text_color: "Error"
            bold: True
            halign: 'center'
            valign: 'middle'
            size_hint: (None, None)
            size: (dp(80), dp(36))
        Widget: # Spacer Kanan

<ResultNoTablesCell@MDBoxLayout>:
    padding: dp(5)
    md_bg_color: app.theme_cls.bg_normal

    MDBoxLayout:
        orientation: 'horizontal'
        Widget: # Spacer Kiri
        MDLabel:
            text: 'Tak Ada Tabel'
            theme_text_color: "Hint"
            halign: 'center'
            valign: 'middle'
            size_hint: (None, None)
            size: (dp(100), dp(36))
        Widget: # Spacer Kanan

# --- LAYAR UTAMA (KONVERSI) ---
<MainScreen>:
    MDBoxLayout:
        orientation: 'vertical'
        padding: dp(20)
        spacing: dp(15)

        MDLabel:
            text: 'Konverter PDF ke Spreadsheet'
            font_style: 'H5'
            bold: True
            halign: 'center'
            size_hint_y: None
            height: self.texture_size[1]

        MDLabel:
            id: status_label
            text: 'Silakan pilih file (maks 5) dan klik Konversi.'
            font_style: 'Subtitle1'
            halign: 'center'
            size_hint_y: None
            height: self.texture_size[1]
            padding_y: dp(10)

        # Toolbar Tombol
        MDBoxLayout:
            id: button_toolbar
            size_hint_y: None
            height: dp(48)
            spacing: dp(10)
            adaptive_height: True

            MDFillRoundFlatIconButton:
                id: select_button
                text: 'Pilih File'
                icon: 'file-upload'
                on_release: root.select_file()
                md_bg_color: app.theme_cls.accent_color
                size_hint_x: 0.33

            MDFillRoundFlatIconButton:
                id: format_spinner_button
                text: 'Pilih Format'
                icon: 'menu-down'
                on_release: root.format_menu.open()
                size_hint_x: 0.33

            MDFillRoundFlatIconButton:
                id: convert_button
                text: 'Konversi Sekarang'
                icon: 'cog-refresh'
                on_release: root.start_conversion()
                size_hint_x: 0.33
                disabled: True # Dinonaktifkan secara default

        # Kontainer Daftar File (Processing)
        MDBoxLayout:
            id: file_list_container
            orientation: 'vertical'
            size_hint_y: None # Awalnya tidak mengambil ruang vertikal
            height: 0       # Awalnya tinggi 0
            spacing: dp(1)
            opacity: 0      # Awalnya transparan

            MDLabel:
                text: "File untuk Dikonversi"
                font_style: 'H6'
                halign: 'left'
                size_hint_y: None
                height: self.texture_size[1]
                padding_y: dp(5)

            # Header Tabel
            MDGridLayout:
                cols: 3
                size_hint_y: None
                height: dp(40)
                md_bg_color: (0.92, 0.92, 0.92, 1) # Abu-abu muda

                MDLabel:
                    text: "Nama File"
                    bold: True
                    halign: "left"
                    valign: "middle"
                    padding_x: dp(5)
                    size_hint_x: 0.6
                MDLabel:
                    text: "Ukuran"
                    bold: True
                    halign: "center"
                    valign: "middle"
                    size_hint_x: 0.15
                MDLabel:
                    text: "Progres"
                    bold: True
                    halign: "center"
                    valign: "middle"
                    size_hint_x: 0.25

            # Scrollable List untuk File
            MDScrollView:
                size_hint_y: 1 # Ambil sisa ruang vertikal di dalam container
                bar_width: dp(10)
                MDGridLayout:
                    id: file_list_grid
                    cols: 3
                    size_hint_y: None # Tinggi mengikuti konten
                    height: self.minimum_height
                    row_default_height: dp(40)
                    row_force_default: True
                    spacing: dp(1)
                    md_bg_color: app.theme_cls.divider_color # Garis antar baris

        # Spacer (mengisi ruang jika file list belum muncul)
        MDBoxLayout:
            id: main_screen_spacer
            size_hint_y: 1 # Awalnya mengambil sisa ruang

# --- LAYAR HASIL (UNDUH) ---
<ResultScreen>:
    MDBoxLayout:
        orientation: 'vertical'
        padding: dp(20)
        spacing: dp(15)

        MDLabel:
            text: "Proses Selesai"
            font_style: 'H5'
            bold: True
            halign: 'center'
            size_hint_y: None
            height: self.texture_size[1]
            padding_y: dp(10)

        MDLabel:
            id: result_status_label
            text: "Silakan unduh hasil Anda."
            font_style: 'Subtitle1'
            halign: 'center'
            size_hint_y: None
            height: self.texture_size[1]
            padding_y: dp(5)

        # Daftar File Hasil
        MDLabel:
            text: "File Hasil Konversi"
            font_style: 'H6'
            halign: 'left'
            size_hint_y: None
            height: self.texture_size[1]
            padding_y: dp(5)

        # ScrollView untuk Daftar Hasil
        MDScrollView:
            size_hint_y: 1 # Ambil sisa ruang vertikal
            bar_width: dp(10)

            MDBoxLayout: # Layout vertikal di dalam ScrollView
                orientation: 'vertical'
                size_hint_y: None # Tinggi mengikuti konten
                height: self.minimum_height
                spacing: 0 # Tidak ada spasi antara header dan grid

                # Header Tabel Hasil
                MDGridLayout:
                    cols: 3
                    size_hint_y: None
                    height: dp(40)
                    md_bg_color: (0.92, 0.92, 0.92, 1) # Abu-abu muda

                    MDLabel:
                        text: "Nama File"
                        bold: True
                        halign: "left"
                        valign: "middle"
                        padding_x: dp(5)
                        size_hint_x: 0.6
                    MDLabel:
                        text: "Ukuran"
                        bold: True
                        halign: "center"
                        valign: "middle"
                        size_hint_x: 0.15
                    MDLabel:
                        text: "Status"
                        bold: True
                        halign: "center"
                        valign: "middle"
                        size_hint_x: 0.25

                # Grid untuk Daftar Hasil
                MDGridLayout:
                    id: result_file_list_grid
                    cols: 3
                    size_hint_y: None # Tinggi mengikuti konten
                    height: self.minimum_height
                    row_default_height: dp(40)
                    row_force_default: True
                    spacing: dp(1)
                    md_bg_color: app.theme_cls.divider_color # Garis antar baris

        # Layout container untuk menengahkan tombol di bawah
        MDBoxLayout:
            orientation: 'horizontal'
            adaptive_height: True # Tinggi mengikuti tombol
            adaptive_width: True  # Lebar mengikuti tombol + spasi
            pos_hint: {'center_x': 0.5} # Tengah horizontal
            spacing: dp(15)
            padding: dp(15), 0 # Padding atas/bawah 0, kiri/kanan 15

            MDFillRoundFlatIconButton:
                text: "Konversi Lagi"
                icon: 'refresh'
                on_release: app.root.go_to_main_screen()
                md_bg_color: get_color_from_hex("#FF9800") # Oranye

            MDFillRoundFlatIconButton:
                id: save_button_result
                text: 'Unduh Semua'
                icon: 'download-multiple'
                on_release: app.root.main_screen.save_result()
                disabled: True # Dinonaktifkan secara default

# --- SCREEN MANAGER (ROOT WIDGET) ---
<RootScreenManager>:
    main_screen: main_screen_id
    result_screen: result_screen_id

    MainScreen:
        id: main_screen_id
        name: 'main'

    ResultScreen:
        id: result_screen_id
        name: 'result'
"""

# --- Definisi Kelas Widget Kustom ---
class GridCellLabel(MDLabel):
    """Sel label dasar untuk grid di MainScreen."""
    pass

class GridCellLabelRight(GridCellLabel):
    """Sel label dengan perataan tengah untuk grid di MainScreen."""
    pass

class WaitingCellWithDelete(MDBoxLayout):
    """Sel kustom yang menampilkan 'Menunggu' dengan tombol X untuk menghapus."""
    file_index = NumericProperty(-1)

class DownloadButtonCell(MDBoxLayout):
    """Widget kustom untuk sel di grid MainScreen yang berisi tombol unduh."""
    file_index = NumericProperty(-1)

class FailedStatusCell(MDBoxLayout):
    """Sel kustom yang menampilkan status 'Gagal' di grid MainScreen."""
    pass

class NoTablesStatusCell(MDBoxLayout):
    """Sel kustom yang menampilkan status 'Tak Ada Tabel' di grid MainScreen."""
    pass

class ProgressTextCell(GridCellLabelRight):
    """Sel label spesifik untuk menampilkan status progres di MainScreen."""
    pass

class ResultCellLabel(MDLabel):
    """Sel label dasar untuk grid di ResultScreen."""
    pass

class ResultCellLabelRight(ResultCellLabel):
    """Sel label dengan perataan tengah untuk grid di ResultScreen."""
    pass

class ResultSuccessCell(MDBoxLayout):
    """Widget kustom untuk sel di grid hasil yang berisi tombol unduh."""
    file_index = NumericProperty(-1)

class ResultFailedCell(MDBoxLayout):
    """Sel kustom yang menampilkan status 'Gagal' di grid ResultScreen."""
    pass

class ResultNoTablesCell(MDBoxLayout):
    """Sel kustom yang menampilkan status 'Tak Ada Tabel' di grid ResultScreen."""
    pass

# --- Muat String KV ---
Builder.load_string(KV)

# --- Fungsi Logika Backend (Pemrosesan Gambar & PDF) ---
def preprocess_image_for_ocr(img_array):
    """Mempersiapkan gambar untuk OCR (grayscale, thresholding, sharpen) agar teks lebih jelas."""
    if len(img_array.shape) == 3:
        gray = cv2.cvtColor(img_array, cv2.COLOR_BGR2GRAY)
    else:
        gray = img_array.copy()
    binary = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 11, 2
    )
    kernel = np.ones((2, 2), np.uint8)
    dilated = cv2.dilate(binary, kernel, iterations=1)
    edge_enhanced = cv2.Laplacian(gray, cv2.CV_8U, ksize=3)
    sharpened = cv2.addWeighted(gray, 1.5, edge_enhanced, -0.5, 0)
    return sharpened

def detect_lines(img_gray, min_length=50, orientation='horizontal', line_type='any'):
    """Mendeteksi garis horizontal atau vertikal menggunakan pemrosesan morfologi dan analisis profil."""
    height, width = img_gray.shape
    binary = cv2.adaptiveThreshold(
        img_gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 11, 2
    )
    if line_type in ['dotted', 'any']:
        if orientation == 'horizontal':
            profile = np.sum(binary, axis=1)
            smooth_profile = gaussian_filter1d(profile, sigma=2)
            lines_img = np.zeros_like(binary)
            peaks, _ = find_peaks(smooth_profile, height=width*0.1, distance=15)
            for peak in peaks:
                if peak > 0 and peak < height:
                    lines_img[peak, :] = 255
        else:
            profile = np.sum(binary, axis=0)
            smooth_profile = gaussian_filter1d(profile, sigma=2)
            lines_img = np.zeros_like(binary)
            peaks, _ = find_peaks(smooth_profile, height=height*0.1, distance=15)
            for peak in peaks:
                if peak > 0 and peak < width:
                    lines_img[:, peak] = 255
        if line_type == 'any':
            if orientation == 'horizontal':
                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (min_length, 1))
            else:
                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, min_length))
            morph_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel, iterations=1)
            combined_lines = cv2.bitwise_or(lines_img, morph_lines)
            return combined_lines
        return lines_img
    else:
        if orientation == 'horizontal':
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (min_length, 1))
        else:
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, min_length))
        lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel, iterations=1)
        return lines

def detect_tables(img_array, line_sensitivity=25, use_grid=True, dotted_line_support=True):
    """Mendeteksi area tabel. Prioritaskan deteksi berbasis grid (garis). Jika gagal, gunakan deteksi berbasis tata letak teks."""
    gray = cv2.cvtColor(img_array, cv2.COLOR_BGR2GRAY)
    if use_grid:
        if dotted_line_support:
            horizontal_lines = detect_lines(gray, min_length=line_sensitivity, orientation='horizontal', line_type='any')
            vertical_lines = detect_lines(gray, min_length=line_sensitivity, orientation='vertical', line_type='any')
        else:
            binary = cv2.adaptiveThreshold(
                gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY_INV, 11, 2
            )
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (line_sensitivity, 1))
            horizontal_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, line_sensitivity))
            vertical_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
        table_grid = cv2.bitwise_or(horizontal_lines, vertical_lines)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        dilated_grid = cv2.dilate(table_grid, kernel, iterations=2)
        contours, _ = cv2.findContours(dilated_grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        table_contours = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            area = w * h
            if area > (img_array.shape[0] * img_array.shape[1]) / 30:
                aspect_ratio = w / float(h)
                if 0.2 < aspect_ratio < 5:
                    table_contours.append((x, y, w, h, horizontal_lines, vertical_lines))
        if table_contours:
            return table_contours
    # Fallback jika deteksi grid gagal
    return text_based_table_detection(img_array, gray)

def text_based_table_detection(img_array, gray):
    """Menganalisis hasil OCR untuk mengelompokkan teks yang berdekatan secara horizontal dan vertikal untuk 'menebak' batas tabel."""
    preprocessed_img = preprocess_image_for_ocr(img_array)
    ocr_result = engine(preprocessed_img)
    if not ocr_result or len(ocr_result[0]) == 0:
        return []
    text_boxes = []
    for box, text, confidence in ocr_result[0]:
        if text.strip():
            x_coords = [p[0] for p in box]
            y_coords = [p[1] for p in box]
            x_min, y_min = min(x_coords), min(y_coords)
            x_max, y_max = max(x_coords), max(y_coords)
            text_boxes.append({
                'box': box, 'text': text, 'confidence': confidence,
                'x': x_min, 'y': y_min, 'width': x_max - x_min, 'height': y_max - y_min,
                'x_center': (x_min + x_max) / 2, 'y_center': (y_min + y_max) / 2
            })
    if not text_boxes:
        return []
    text_boxes_sorted_y = sorted(text_boxes, key=lambda b: b['y'])
    avg_height = sum(box['height'] for box in text_boxes) / len(text_boxes)
    row_threshold = max(avg_height * 1.5, 15)
    rows = []
    current_row = [text_boxes_sorted_y[0]]
    current_y = text_boxes_sorted_y[0]['y_center']
    for box in text_boxes_sorted_y[1:]:
        if abs(box['y_center'] - current_y) <= row_threshold:
            current_row.append(box)
        else:
            rows.append(current_row)
            current_row = [box]
            current_y = box['y_center']
    if current_row:
        rows.append(current_row)
    if len(rows) < 2:
        return []
    table_candidates = []
    min_rows_for_table = 3
    for i in range(len(rows) - min_rows_for_table + 1):
        for k in range(len(rows) - i - min_rows_for_table + 1):
            potential_table_rows = rows[i:i+min_rows_for_table+k]
            row_lengths = [len(row) for row in potential_table_rows]
            if max(row_lengths) - min(row_lengths) <= 1:
                all_boxes = [box for row in potential_table_rows for box in row]
                x_min = min(box['x'] for box in all_boxes)
                y_min = min(box['y'] for box in all_boxes)
                x_max = max(box['x'] + box['width'] for box in all_boxes)
                y_max = max(box['y'] + box['height'] for box in all_boxes)
                x_min = max(0, x_min - 10)
                y_min = max(0, y_min - 10)
                x_max = min(img_array.shape[1], x_max + 10)
                y_max = min(img_array.shape[0], y_max + 10)
                empty_lines = np.zeros_like(gray)
                table_candidates.append((int(x_min), int(y_min), int(x_max - x_min), int(y_max - y_min), empty_lines, empty_lines))
    return table_candidates

def detect_cells_from_text(text_items, table_width, table_height):
    """Mengelompokkan koordinat X dan Y dari teks untuk menyimpulkan batas baris dan kolom (grid imajiner)."""
    if not text_items:
        return None, None
    avg_height = sum(item['height'] for item in text_items) / len(text_items)
    avg_width = sum(item['width'] for item in text_items) / len(text_items)
    sorted_by_y = sorted(text_items, key=lambda t: t['y_center'])
    y_centers = [t['y_center'] for t in sorted_by_y]
    y_threshold = max(avg_height * 0.8, 10)
    def cluster_coordinates(coords, threshold):
        clusters = []
        if not coords:
            return clusters
        current_cluster = [coords[0]]
        current_value = coords[0]
        for val in coords[1:]:
            if abs(val - current_value) <= threshold:
                current_cluster.append(val)
            else:
                if current_cluster:
                    clusters.append(sum(current_cluster) / len(current_cluster))
                current_cluster = [val]
                current_value = val
        if current_cluster:
            clusters.append(sum(current_cluster) / len(current_cluster))
        return clusters
    row_centers = cluster_coordinates(y_centers, y_threshold)
    row_boundaries = [0]
    for i in range(len(row_centers) - 1):
        boundary = int((row_centers[i] + row_centers[i+1]) / 2)
        row_boundaries.append(boundary)
    row_boundaries.append(int(table_height))
    sorted_by_x = sorted(text_items, key=lambda t: t['x_center'])
    x_centers = [t['x_center'] for t in sorted_by_x]
    x_threshold = max(avg_width * 0.8, 20)
    col_centers = cluster_coordinates(x_centers, x_threshold)
    col_boundaries = [0]
    for i in range(len(col_centers) - 1):
        boundary = int((col_centers[i] + col_centers[i+1]) / 2)
        col_boundaries.append(boundary)
    col_boundaries.append(int(table_width))
    return row_boundaries, col_boundaries

def process_table_ocr(ocr_results, table_rect, use_inferred_grid=True):
    """Mengekstrak teks di dalam batas tabel dan memetakannya ke struktur sel (baris/kolom) berdasarkan grid yang disimpulkan."""
    x_table, y_table, w_table, h_table = table_rect[:4]
    if not ocr_results or len(ocr_results[0]) == 0:
        return None
    table_texts = []
    for box, text, confidence in ocr_results[0]:
        x_center = sum([p[0] for p in box]) / 4
        y_center = sum([p[1] for p in box]) / 4
        if ((x_table - 5 <= x_center <= x_table + w_table + 5) and
            (y_table - 5 <= y_center <= y_table + h_table + 5)):
            x_min = min([p[0] for p in box])
            y_min = min([p[1] for p in box])
            if text.strip():
                table_texts.append({
                    'text': text.strip(), 'confidence': confidence,
                    'x': x_min - x_table, 'y': y_min - y_table,
                    'width': max([p[0] for p in box]) - x_min, 'height': max([p[1] for p in box]) - y_min,
                    'x_center': x_center - x_table, 'y_center': y_center - y_table, 'original_box': box
                })
    if not table_texts:
        return None
    if use_inferred_grid:
        row_boundaries, col_boundaries = detect_cells_from_text(table_texts, w_table, h_table)
        if row_boundaries and col_boundaries and len(row_boundaries) > 1 and len(col_boundaries) > 1:
            num_rows = len(row_boundaries) - 1
            num_cols = len(col_boundaries) - 1
            table_data = [[''] * num_cols for _ in range(num_rows)]
            for text_item in table_texts:
                row_idx = None
                for i in range(len(row_boundaries) - 1):
                    if row_boundaries[i] <= text_item['y_center'] < row_boundaries[i+1]:
                        row_idx = i
                        break
                col_idx = None
                for i in range(len(col_boundaries) - 1):
                    if col_boundaries[i] <= text_item['x_center'] < col_boundaries[i+1]:
                        col_idx = i
                        break
                if row_idx is not None and col_idx is not None:
                    if row_idx < num_rows and col_idx < num_cols:
                        if table_data[row_idx][col_idx]:
                            table_data[row_idx][col_idx] += ' ' + text_item['text']
                        else:
                            table_data[row_idx][col_idx] = text_item['text']
            return table_data
    # Fallback jika grid imajiner gagal
    return reconstruct_table(table_texts, w_table, h_table)

def reconstruct_table(texts, table_width, table_height):
    """Metode alternatif untuk membangun tabel jika grid gagal, hanya berdasarkan pengelompokan Y (baris) dan X (kolom)."""
    if not texts:
        return None
    texts_sorted_by_y = sorted(texts, key=lambda t: t['y'])
    avg_text_height = sum(t['height'] for t in texts) / len(texts)
    row_threshold = max(avg_text_height * 0.8, 12)
    rows = []
    if not texts_sorted_by_y:
        return []
    current_row = [texts_sorted_by_y[0]]
    row_y = texts_sorted_by_y[0]['y_center']
    for t in texts_sorted_by_y[1:]:
        if abs(t['y_center'] - row_y) <= row_threshold:
            current_row.append(t)
        else:
            if current_row:
                rows.append(sorted(current_row, key=lambda t: t['x']))
            current_row = [t]
            row_y = t['y_center']
    if current_row:
        rows.append(sorted(current_row, key=lambda t: t['x']))
    all_x_centers = [t['x_center'] for row in rows for t in row]
    def cluster_x_centers(x_centers):
        if not x_centers:
            return []
        avg_text_width = sum(t['width'] for t in texts) / len(texts)
        col_threshold = max(avg_text_width * 1.5, 20)
        x_centers = sorted(x_centers)
        clusters = [[x_centers[0]]]
        for x in x_centers[1:]:
            if abs(x - clusters[-1][-1]) <= col_threshold:
                clusters[-1].append(x)
            else:
                clusters.append([x])
        return [sum(cluster) / len(cluster) for cluster in clusters]
    col_centers = cluster_x_centers(all_x_centers)
    num_cols = len(col_centers)
    if num_cols == 0:
        num_cols = 1
        col_centers = [table_width / 2]
    table_data = []
    for row_texts in rows:
        row_data = [''] * num_cols
        for text in row_texts:
            col_idx = min(range(num_cols),
                          key=lambda i: abs(text['x_center'] - col_centers[i]))
            if row_data[col_idx]:
                row_data[col_idx] += ' ' + text['text']
            else:
                row_data[col_idx] = text['text']
        table_data.append(row_data)
    return table_data

def create_tables_export(tables, output_format='xlsx', sheet_names=None):
    """Mengambil data tabel (list of DataFrame) dan mengekspornya ke format file yang diminta (XLSX, CSV, ODS) dalam bentuk bytes."""
    output = io.BytesIO()
    def auto_adjust_excel_column_widths(writer, sheet_name, dataframe):
        worksheet = writer.book[sheet_name]
        for idx, col_name in enumerate(dataframe.columns):
            max_length = len(str(col_name))
            for cell in dataframe[col_name].astype(str):
                if len(cell) > max_length:
                    max_length = len(cell)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width

    file_ext = output_format
    mime_type = "application/octet-stream"

    if output_format == 'xlsx':
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                if tables:
                    for i, df in enumerate(tables):
                        safe_sheet_name = "".join(c for c in sheet_names[i] if c.isalnum() or c in (' ', '_')).rstrip()[:31] if sheet_names and i < len(sheet_names) else f"Tabel_{i+1}"
                        temp_name = safe_sheet_name
                        count = 1
                        while temp_name in writer.book.sheetnames:
                            temp_name = f"{safe_sheet_name}_{count}"[:31]
                            count += 1
                        safe_sheet_name = temp_name
                        df.to_excel(writer, sheet_name=safe_sheet_name, index=False, header=False)
                        auto_adjust_excel_column_widths(writer, safe_sheet_name, df)
                else:
                    wb = Workbook()
                    wb.active.title = "Tidak_Ada_Tabel"
                    wb.save(output)
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            file_ext = "xlsx"
        except Exception as e:
            print(f"Gagal membuat XLSX: {e}")
            output = io.BytesIO()

    elif output_format == 'csv':
        if not tables:
            output.write(b'')
        else:
             # Hanya ekspor tabel pertama jika CSV
             output.write(tables[0].to_csv(index=False, header=False, encoding='utf-8-sig').encode('utf-8-sig'))
        mime_type = "text/csv"
        file_ext = "csv"

    elif output_format == 'ods':
        try:
            with pd.ExcelWriter(output, engine='odf') as writer:
                if tables:
                    for i, df in enumerate(tables):
                        sheet_name = "".join(c for c in sheet_names[i] if c.isalnum() or c in (' ', '_')).rstrip()[:31] if sheet_names and i < len(sheet_names) else f"Tabel_{i+1}"
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            mime_type = "application/vnd.oasis.opendocument.spreadsheet"
            file_ext = "ods"
        except Exception as e:
            print(f"Gagal membuat ODS: {e}")
            output = io.BytesIO()

    output.seek(0)
    return output.getvalue(), mime_type, file_ext

def process_image_for_tables(img_array):
    """Orkestrator utama untuk satu gambar: menjalankan deteksi tabel, lalu OCR, lalu mem-parsing tabel."""
    use_grid = True
    line_sensitivity = 25
    dotted_line_support = True
    small_text_enhancement = True
    if small_text_enhancement:
        preprocessed_img = preprocess_image_for_ocr(img_array)
        ocr_result = engine(preprocessed_img)
    else:
        ocr_result = engine(img_array)
    table_rects = detect_tables(img_array, line_sensitivity, use_grid, dotted_line_support)
    if not table_rects:
        # Jika tidak ada tabel terdeteksi, kembalikan semua teks sebagai satu kolom
        if isinstance(ocr_result, tuple) and len(ocr_result) >= 1 and ocr_result[0]:
            all_texts = [item[1] for item in ocr_result[0]]
            if all_texts:
                return [pd.DataFrame({"Teks": all_texts})]
        return []
    tables = []
    for i, table_rect in enumerate(table_rects):
        table_data = process_table_ocr(ocr_result, table_rect, use_inferred_grid=True)
        if table_data:
            df = pd.DataFrame(table_data)
            tables.append(df)
    return tables


def process_pdf(pdf_path, progress_callback=None):
    """Mengkonversi setiap halaman PDF menjadi gambar, lalu memproses setiap gambar untuk tabel."""
    all_tables = []
    sheet_names = []

    try:
        pdf_info = pdf2image.pdfinfo_from_path(pdf_path)
        total_pages = pdf_info.get('Pages', 0)
        if total_pages == 0:
             raise ValueError("Tidak dapat menentukan jumlah halaman.")
    except Exception as e:
        print(f"Gagal mendapatkan info PDF untuk {pdf_path}: {e}")
        if progress_callback:
            progress_callback(0, 1, error=True, message=f"Gagal Info")
        return [], []

    for i in range(1, total_pages + 1):
        if progress_callback:
            # Kirim progres ke UI
            progress_callback(i, total_pages)

        try:
            page_image = pdf2image.convert_from_path(
                pdf_path,
                dpi=300,
                first_page=i,
                last_page=i,
                timeout=3600
            )[0]

            img_array = np.array(page_image.convert('RGB'))
            page_tables = process_image_for_tables(img_array)

            for j, df in enumerate(page_tables):
                all_tables.append(df)
                sheet_names.append(f"Halaman {i}")

            # Bersihkan memori
            del page_image
            del img_array
            gc.collect()
        except Exception as e:
            print(f"Gagal memproses halaman {i} dari {pdf_path}: {e}")
            if progress_callback:
                 progress_callback(i, total_pages, error=True, message=f"Halaman {i} Gagal")
            continue

    return all_tables, sheet_names

def process_image(image_path):
    """Wrapper sederhana untuk memproses file gambar tunggal."""
    try:
        image = Image.open(image_path)
        img_array = np.array(image.convert('RGB'))
        tables = process_image_for_tables(img_array)
        del image
        del img_array
        gc.collect()
        return tables
    except Exception as e:
        print(f"Gagal memproses gambar {image_path}: {e}")
        return []


# --- Definisi Layar ---
class ResultScreen(MDScreen):
    """Layar untuk menampilkan hasil konversi."""
    pass # Logika UI ada di KV

class MainScreen(MDScreen):
    """Layar utama aplikasi untuk memilih file dan memulai konversi."""
    max_files = 5 # Batas jumlah file

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # Inisialisasi variabel state aplikasi
        self.extracted_tables = {} # {index: [df1, df2], ...}
        self.sheet_names = {}      # {index: ["sheet1", "sheet2"], ...}
        self.output_data = {}      # {index: bytes}
        self.file_list = []        # List of dicts: {'index', 'path', 'name', ..., 'status', 'progress_widget'}
        self.current_processing_index = -1
        self.is_processing = False
        self._dropped_files = []
        self._drop_event = None
        self.selected_format = None # Format output yang dipilih (misal 'XLSX')
        self.format_menu = None     # Objek MDDropdownMenu
        self.dialog = None          # Objek MDDialog untuk alert

        # Ikat fungsi on_file_drop ke window
        Window.bind(on_dropfile=self.on_file_drop)
        # Buat menu format setelah frame pertama
        Clock.schedule_once(self.create_format_menu)

    def show_alert(self, title, text):
        """Menampilkan dialog peringatan modal."""
        if self.dialog and self.dialog.is_open:
            self.dialog.dismiss()

        # Buat konten dialog secara dinamis
        text_label = MDLabel(
            text=text,
            theme_text_color="Primary",
            halign='center',
            valign='top',
            size_hint_y=None,
            markup=True
        )
        text_label.bind(
            width=lambda instance, value: setattr(instance, 'text_size', (value * 0.9, None)) # Agar pas
        )

        ok_button = MDFillRoundFlatButton(
            text="OK",
            size_hint=(None, None),
            size=(dp(100), dp(40)),
            pos_hint={'center_x': 0.5},
            on_release=lambda x: self.dialog.dismiss()
        )

        content_box = MDBoxLayout(
            orientation='vertical',
            adaptive_height=True,
            padding=dp(20),
            spacing=dp(20)
        )
        content_box.add_widget(text_label)
        content_box.add_widget(ok_button)

        self.dialog = MDDialog(
            title=title,
            type="custom",
            content_cls=content_box,
            size_hint=(0.8, None),
            auto_dismiss=False
        )

        # Atur title alignment (perlu dijadwalkan agar ID tersedia)
        def set_title_center(*args):
             try:
                 self.dialog.ids.title.halign = 'center'
             except Exception as e:
                  print(f"Warning: could not center dialog title: {e}")

        Clock.schedule_once(set_title_center, 0.1)
        self.dialog.open()


    def create_format_menu(self, *args):
        """Membangun item menu dropdown untuk pilihan format."""
        menu_items = [
            {"text": "XLSX", "viewclass": "OneLineListItem", "height": dp(48),
             "on_release": lambda x="XLSX": self.set_format(x)},
            {"text": "CSV", "viewclass": "OneLineListItem", "height": dp(48),
             "on_release": lambda x="CSV": self.set_format(x)},
            {"text": "ODS", "viewclass": "OneLineListItem", "height": dp(48),
             "on_release": lambda x="ODS": self.set_format(x)},
        ]
        # Pastikan id 'format_spinner_button' ada sebelum membuat menu
        if self.ids and 'format_spinner_button' in self.ids:
            self.format_menu = MDDropdownMenu(
                caller=self.ids.format_spinner_button,
                items=menu_items,
                width_mult=3,
            )
        else:
             print("Error: ID 'format_spinner_button' tidak ditemukan saat membuat menu.")


    def set_format(self, format_text):
        """Dipanggil saat format dipilih dari dropdown."""
        self.ids.format_spinner_button.text = format_text
        self.selected_format = format_text.lower() # Simpan dalam lowercase
        self.format_menu.dismiss()
        # Aktifkan tombol konversi hanya jika ada file dan format dipilih
        self.ids.convert_button.disabled = not (self.file_list and self.selected_format)


    def _clear_results(self):
        """Membersihkan data hasil konversi sebelumnya."""
        self.extracted_tables.clear()
        self.sheet_names.clear()
        self.output_data.clear()
        for f_info in self.file_list:
            f_info['output_data'] = None
            f_info['tables'] = None
            f_info['sheet_names'] = None
            f_info['status'] = 'pending' # Reset status
        gc.collect()
        print("Hasil ekstraksi sebelumnya telah dibersihkan.")


    def get_file_size_str(self, file_path):
        """Mengubah ukuran file (bytes) menjadi string yang mudah dibaca."""
        try:
            size_bytes = os.path.getsize(file_path)
            if size_bytes < 1024: return f"{size_bytes} B"
            elif size_bytes < 1024**2: return f"{size_bytes/1024:.1f} KB"
            elif size_bytes < 1024**3: return f"{size_bytes/1024**2:.1f} MB"
            else: return f"{size_bytes/1024**3:.1f} GB"
        except Exception: return "N/A"


    @mainthread
    def reset_ui_state(self):
        """Mereset UI ke kondisi awal."""
        self.ids.status_label.text = f"Silakan pilih file (maks {self.max_files}) dan klik Konversi."
        self.ids.file_list_grid.clear_widgets()
        self.file_list = []

        # Sembunyikan container file list dan tampilkan spacer
        self.ids.file_list_container.opacity = 0
        self.ids.file_list_container.size_hint_y = None
        self.ids.file_list_container.height = 0
        self.ids.main_screen_spacer.size_hint_y = 1
        self.ids.main_screen_spacer.height = dp(1) # Beri tinggi minimal agar spacer efektif

        # Tampilkan toolbar tombol
        self.ids.button_toolbar.opacity = 1
        self.ids.button_toolbar.size_hint_y = None
        self.ids.button_toolbar.height = dp(48)

        # Reset tombol
        self.ids.select_button.disabled = False
        self.ids.convert_button.disabled = True # Nonaktifkan tombol konversi
        self.ids.format_spinner_button.text = "Pilih Format"
        self.selected_format = None

        # Reset state internal
        self.is_processing = False
        self.current_processing_index = -1
        self._clear_results() # Hapus data hasil sebelumnya


    def select_file(self):
        """Membuka dialog pilih file."""
        if self.is_processing: return

        if len(self.file_list) >= self.max_files:
            self.show_alert("Batas File Tercapai", f"Anda sudah memilih {self.max_files} file. Batas maksimal adalah {self.max_files} file.")
            return

        try:
            filechooser.open_file(
                on_selection=self.handle_selection,
                filters=[("PDF & Gambar", "*.pdf", "*.jpeg", "*.jpg", "*.png")],
                multiple=True
            )
        except Exception as e:
             print(f"Gagal membuka file chooser: {e}")
             self.show_alert("Error", "Gagal membuka dialog pilih file.")

    def handle_selection(self, selection):
        """Menangani file yang dipilih dari dialog atau drag-and-drop."""
        if not selection or self.is_processing:
            return

        current_file_count = len(self.file_list)
        allowed_new_count = self.max_files - current_file_count

        if allowed_new_count <= 0 and selection:
            self.show_alert("Batas File Tercapai", f"Anda sudah memilih {self.max_files} file. Tidak dapat menambah lagi.")
            return

        existing_paths = {f['path'] for f in self.file_list}
        added_count = 0
        rejected_count = 0
        new_paths_to_add = []

        for path in selection:
            if added_count >= allowed_new_count:
                rejected_count += 1
                continue
            # Filter duplikat dan tipe file
            allowed_extensions = ('.pdf', '.jpeg', '.jpg', '.png')
            if path not in existing_paths and path.lower().endswith(allowed_extensions):
                new_paths_to_add.append(path)
                existing_paths.add(path) # Tambahkan ke set agar tidak duplikat dalam batch yang sama
                added_count += 1
            else:
                rejected_count += 1

        if rejected_count > 0:
            reason = "batas file tercapai" if added_count >= allowed_new_count else "file duplikat atau format tidak didukung"
            self.show_alert(
                "Info Pemilihan File",
                f"{added_count} file berhasil ditambahkan.\n"
                f"{rejected_count} file ditolak ({reason})."
            )

        if new_paths_to_add:
            self._append_files_to_ui(new_paths_to_add)


    def on_file_drop(self, window, file_path_bytes):
        """Handler event saat file di-drag ke jendela aplikasi."""
        if self.is_processing: return

        try:
            file_path_str = file_path_bytes.decode('utf-8')
            # Tambahkan file yang valid ke list sementara
            allowed_extensions = ('.pdf', '.jpeg', '.jpg', '.png')
            if file_path_str.lower().endswith(allowed_extensions):
                 # Cek batas sebelum menambahkan ke _dropped_files
                 if len(self.file_list) + len(self._dropped_files) < self.max_files:
                    self._dropped_files.append(file_path_str)
                 elif not self._drop_event: # Tampilkan alert hanya sekali per drop batch
                      Clock.schedule_once(lambda dt: self.show_alert("Batas File Tercapai", f"Batas maksimal {self.max_files} file akan tercapai. Beberapa file yang di-drop mungkin ditolak."), 0.1)

        except Exception as e:
            print(f"Gagal decode path file drag-drop: {e}")
            return

        # Jadwalkan pemrosesan batch setelah jeda singkat
        if self._drop_event:
            self._drop_event.cancel()
        self._drop_event = Clock.schedule_once(self.process_dropped_files, 0.2)


    def process_dropped_files(self, dt):
        """Memproses batch file yang di-drop."""
        self._drop_event = None # Reset event schedule
        if not self._dropped_files or self.is_processing:
            self._dropped_files.clear()
            return

        files_to_process = list(self._dropped_files) # Salin list
        self._dropped_files.clear() # Kosongkan list asli
        self.handle_selection(files_to_process) # Proses menggunakan logika yang sama


    @mainthread
    def _append_files_to_ui(self, new_paths):
        """Menambahkan representasi visual file baru ke grid UI."""
        if not new_paths: return

        # Tampilkan container file list jika perlu
        if self.ids.file_list_container.opacity == 0:
            self.ids.file_list_container.opacity = 1
            self.ids.file_list_container.size_hint_y = 1
            self.ids.file_list_container.height = dp(1) # Beri tinggi awal agar size_hint_y=1 bekerja
            self.ids.main_screen_spacer.size_hint_y = None
            self.ids.main_screen_spacer.height = 0

        grid = self.ids.file_list_grid
        start_index = len(self.file_list)

        for i, path in enumerate(new_paths):
            current_index = start_index + i
            try:
                file_name = os.path.basename(path)
                file_size = self.get_file_size_str(path)

                # Widget untuk nama file
                label_name = GridCellLabel(text=file_name, size_hint_x=0.6)
                # Widget untuk ukuran file
                label_size = GridCellLabelRight(text=file_size, size_hint_x=0.15)
                # Widget awal untuk status (Menunggu + tombol hapus)
                progress_widget = WaitingCellWithDelete(file_index=current_index, size_hint_x=0.25)

                grid.add_widget(label_name)
                grid.add_widget(label_size)
                grid.add_widget(progress_widget)

                # Tambahkan info file ke list internal
                self.file_list.append({
                    'index': current_index, 'path': path, 'name': file_name,
                    'original_name': os.path.splitext(file_name)[0], 'size': file_size,
                    'status': 'pending', # Status awal
                    'progress_widget': progress_widget, # Simpan referensi widget status
                    # Data hasil akan diisi nanti
                    'output_data': None, 'tables': None, 'sheet_names': None, 'file_ext': None
                })

            except Exception as e:
                print(f"Gagal menambahkan UI untuk file {path}: {e}")
                # Tambahkan baris error jika gagal
                grid.add_widget(GridCellLabel(text=f"Error: {os.path.basename(path)}", size_hint_x=0.6, color=(0.8,0,0,1)))
                grid.add_widget(GridCellLabelRight(text="N/A", size_hint_x=0.15))
                grid.add_widget(GridCellLabelRight(text="Gagal Muat", size_hint_x=0.25, color=(0.8,0,0,1)))

        # Update label status
        self.ids.status_label.text = f"{len(self.file_list)} file siap dikonversi."
        # Aktifkan tombol konversi jika format sudah dipilih
        self.ids.convert_button.disabled = not self.selected_format


    def start_conversion(self):
        """Memulai proses konversi untuk semua file yang 'pending'."""
        if self.is_processing:
            self.show_alert("Info", "Proses konversi sedang berjalan.")
            return
        if not self.file_list:
            self.show_alert("Peringatan", "Tidak ada file yang dipilih untuk dikonversi.")
            return
        if not self.selected_format:
             self.show_alert("Peringatan", "Silakan pilih format output terlebih dahulu.")
             return

        # Cek apakah ada file yang perlu direset statusnya (jika semua sudah selesai/gagal)
        if not any(f['status'] == 'pending' for f in self.file_list):
            self.ids.status_label.text = "Mereset status file untuk konversi ulang..."
            print("Mereset status file untuk konversi ulang...")
            self._clear_results() # Hapus data lama dan reset status ke 'pending'
            # Reset UI widget progres untuk semua file
            grid = self.ids.file_list_grid
            grid.clear_widgets() # Hapus semua widget lama
            temp_paths = [f['path'] for f in self.file_list] # Ambil path
            self.file_list = [] # Kosongkan list file
            self._append_files_to_ui(temp_paths) # Tambahkan kembali ke UI dengan status 'pending'
            Clock.schedule_once(lambda dt: self.start_conversion(), 0.5) # Coba mulai lagi setelah UI di-refresh
            return


        self.is_processing = True
        self._clear_results() # Pastikan hasil lama bersih & status 'pending'

        # Kunci UI: Sembunyikan toolbar tombol
        self.ids.button_toolbar.opacity = 0
        self.ids.button_toolbar.size_hint_y = None
        self.ids.button_toolbar.height = 0
        # self.ids.select_button.disabled = True # Sebenarnya sudah tersembunyi

        self.ids.status_label.text = "Memulai konversi..."

        # Ganti widget 'WaitingCellWithDelete' menjadi 'ProgressTextCell' "Menunggu"
        grid = self.ids.file_list_grid
        for i, f_info in enumerate(self.file_list):
            if f_info['status'] == 'pending':
                current_widget = f_info.get('progress_widget')
                # Cari widget di children grid jika referensi hilang (berdasarkan index relatif dari belakang)
                widget_list_index = (len(self.file_list) - 1 - i) * 3 # Index child ke-0 dari triplet terakhir
                if not current_widget or current_widget not in grid.children:
                     try:
                         current_widget = grid.children[widget_list_index]
                     except IndexError:
                          print(f"Error: Tidak dapat menemukan widget progress untuk index {i} saat memulai.")
                          continue # Lanjut ke file berikutnya jika widget tidak ada

                if isinstance(current_widget, WaitingCellWithDelete):
                    try:
                        widget_visual_index = grid.children.index(current_widget)
                        grid.remove_widget(current_widget)

                        new_progress_widget = ProgressTextCell(
                            text="Menunggu", size_hint_x=0.25, color=(0.5, 0.5, 0.5, 1)
                        )
                        grid.add_widget(new_progress_widget, index=widget_visual_index)
                        f_info['progress_widget'] = new_progress_widget # Update referensi
                    except Exception as e:
                        print(f"Error mengganti widget 'Menunggu' untuk index {i}: {e}")

        # Cari file 'pending' pertama untuk memulai
        first_pending_index = -1
        for i, f_info in enumerate(self.file_list):
            if f_info['status'] == 'pending':
                first_pending_index = i
                break

        if first_pending_index != -1:
            self.current_processing_index = first_pending_index
            self._process_next_file() # Mulai proses dari file pertama
        else:
            # Seharusnya tidak terjadi jika logika reset di atas benar
            self.is_processing = False
            self.ids.status_label.text = "Tidak ada file yang perlu diproses."
            self.reset_ui_state() # Kembali ke state awal jika tidak ada yg diproses


    @mainthread
    def remove_file_from_list(self, file_index):
        """Menghapus file dari daftar SEBELUM konversi dimulai."""
        if self.is_processing:
            self.show_alert("Info", "Tidak dapat menghapus file saat proses konversi berjalan.")
            return

        if not (0 <= file_index < len(self.file_list)):
            print(f"Error: Index {file_index} tidak valid untuk dihapus.")
            return

        file_info = self.file_list[file_index]

        # Hanya bisa dihapus jika status 'pending' dan widget-nya WaitingCellWithDelete
        if file_info['status'] != 'pending' or not isinstance(file_info.get('progress_widget'), WaitingCellWithDelete):
            # self.show_alert("Info", "File ini tidak dapat dihapus saat ini.")
            print(f"File index {file_index} tidak dalam status 'pending' atau widget salah.")
            return

        grid = self.ids.file_list_grid
        widgets_to_remove = []
        try:
            # Cari 3 widget terkait berdasarkan index file (dari belakang grid.children)
            # Urutan di children: [progress_n, size_n, name_n, progress_n-1, ...]
            base_idx = (len(self.file_list) - 1 - file_index) * 3
            if base_idx + 2 < len(grid.children):
                 widgets_to_remove.append(grid.children[base_idx])     # progress
                 widgets_to_remove.append(grid.children[base_idx + 1]) # size
                 widgets_to_remove.append(grid.children[base_idx + 2]) # name
            else:
                 raise IndexError("Indeks widget di luar jangkauan.")

        except Exception as e:
            print(f"Error saat mencari widget untuk dihapus (index {file_index}): {e}")
            self.show_alert("Error Internal", "Gagal menghapus file dari tampilan.")
            return

        # Hapus widget dari grid
        for widget in widgets_to_remove:
            grid.remove_widget(widget)

        # Hapus data file dari list
        removed_file_name = self.file_list.pop(file_index)['name']

        # Re-index file yang tersisa dan update file_index di widget WaitingCellWithDelete
        for i, f_info in enumerate(self.file_list):
            f_info['index'] = i
            if isinstance(f_info.get('progress_widget'), WaitingCellWithDelete):
                f_info['progress_widget'].file_index = i

        print(f"File '{removed_file_name}' (index asli {file_index}) berhasil dihapus.")

        # Update status label atau sembunyikan list jika kosong
        total_files = len(self.file_list)
        if total_files > 0:
            self.ids.status_label.text = f"{total_files} file siap dikonversi."
            self.ids.convert_button.disabled = not self.selected_format # Enable/disable convert btn
        else:
            self.reset_ui_state() # Kembali ke tampilan awal jika tidak ada file


    @mainthread
    def reset_progress_widget(self, file_index):
        """Mereset widget status file ke 'WaitingCellWithDelete' (jarang digunakan langsung)."""
        if not (0 <= file_index < len(self.file_list)): return

        file_info = self.file_list[file_index]
        grid = self.ids.file_list_grid
        current_widget = file_info.get('progress_widget')

        # Cari widget jika referensi hilang
        widget_list_index = (len(self.file_list) - 1 - file_index) * 3
        if not current_widget or current_widget not in grid.children:
            try:
                current_widget = grid.children[widget_list_index]
            except IndexError:
                print(f"Error: Tidak dapat menemukan widget progress untuk reset (index {file_index}).")
                return

        # Hanya reset jika bukan WaitingCellWithDelete
        if not isinstance(current_widget, WaitingCellWithDelete):
            try:
                widget_visual_index = grid.children.index(current_widget)
                grid.remove_widget(current_widget)

                new_widget = WaitingCellWithDelete(file_index=file_index, size_hint_x=0.25)
                grid.add_widget(new_widget, index=widget_visual_index)

                file_info['progress_widget'] = new_widget
                file_info['status'] = 'pending' # Pastikan status juga direset
                print(f"Widget progress untuk index {file_index} direset.")
            except Exception as e:
                 print(f"Error saat mereset widget progress index {file_index}: {e}")
        else:
             file_info['status'] = 'pending' # Pastikan status pending


    def _process_next_file(self):
        """Menemukan file 'pending' berikutnya dan memulai thread pemrosesan."""
        next_index_to_process = -1
        # Cari file berikutnya yang masih 'pending', mulai dari index saat ini
        for i in range(self.current_processing_index, len(self.file_list)):
            if self.file_list[i]['status'] == 'pending':
                next_index_to_process = i
                break

        if next_index_to_process != -1:
            # File 'pending' ditemukan
            self.current_processing_index = next_index_to_process
            file_info = self.file_list[self.current_processing_index]
            file_info['status'] = 'processing' # Tandai sebagai sedang diproses
            self.ids.status_label.text = f"Memproses file {self.current_processing_index + 1}/{len(self.file_list)}: {file_info['name']}..."

            # Update widget progres menjadi "Memproses..."
            progress_widget = file_info.get('progress_widget')
            if progress_widget and hasattr(progress_widget, 'text'):
                progress_widget.text = "Memproses..."
                progress_widget.color = (0, 0, 0, 1) # Hitam

            # Mulai thread baru untuk pemrosesan file ini
            print(f"Memulai thread untuk file: {file_info['name']} (index {file_info['index']})")
            thread = threading.Thread(
                target=self._run_single_conversion_thread,
                args=(file_info['path'], file_info['index']),
                daemon=True # Agar thread berhenti jika aplikasi ditutup
            )
            thread.start()

        else:
            # Tidak ada file 'pending' lagi, semua selesai diproses
            self.is_processing = False
            self.current_processing_index = -1 # Reset index
            self.ids.status_label.text = "Semua file selesai diproses."
            print("Semua file telah diproses.")

            # Siapkan dan pindah ke layar hasil
            self._prepare_and_go_to_result_screen()


    @mainthread
    def _update_progress_label(self, file_index, current_page, total_pages, error=False, message=""):
        """Callback dari thread PDF untuk memperbarui label progres di UI."""
        if not (0 <= file_index < len(self.file_list)): return

        file_info = self.file_list[file_index]
        # Hanya update jika status masih 'processing'
        if file_info['status'] != 'processing': return

        progress_widget = file_info.get('progress_widget')
        if progress_widget and hasattr(progress_widget, 'text'):
            if error:
                progress_widget.text = message if message else "Error Halaman"
                progress_widget.color = (0.8, 0.2, 0.2, 1) # Merah
            elif file_info['path'].lower().endswith('.pdf'):
                progress_widget.text = f"Halaman {current_page}/{total_pages}"
                progress_widget.color = (0, 0, 0, 1) # Hitam
            else: # Untuk gambar (biasanya hanya sekali update)
                progress_widget.text = "Memproses..."
                progress_widget.color = (0, 0, 0, 1) # Hitam


    def _run_single_conversion_thread(self, file_path, file_index):
        """Fungsi yang berjalan di thread terpisah untuk memproses satu file."""
        tables = None
        sheet_names = None
        error_object = None
        print(f"Thread {file_index}: Memulai pemrosesan {os.path.basename(file_path)}")

        try:
            file_extension = os.path.splitext(file_path)[-1].lower()
            if file_extension == '.pdf':
                # Sediakan callback untuk update progres
                progress_callback = lambda current, total, error=False, message="": \
                    self._update_progress_label(file_index, current, total, error, message)
                tables, sheet_names = process_pdf(file_path, progress_callback)
            elif file_extension in ['.jpeg', '.jpg', '.png']:
                 # Update progress sekali untuk gambar
                 self._update_progress_label(file_index, 1, 1)
                 tables = process_image(file_path)
                 # Buat sheet names default untuk gambar
                 sheet_names = [f"Tabel_{i+1}" for i in range(len(tables or []))]
            else:
                 raise ValueError(f"Format file tidak didukung: {file_extension}")

            print(f"Thread {file_index}: Selesai pemrosesan. Hasil tabel: {len(tables) if tables else 0}")

        except Exception as e:
            error_object = e
            print(f"Thread {file_index}: ERROR saat memproses file: {e}")
            self._update_progress_label(file_index, 0, 1, error=True, message="Gagal Total")

        # Kirim hasil (atau error) kembali ke main thread
        Clock.schedule_once(lambda dt: self.on_single_conversion_result(file_index, tables, sheet_names, error_object))


    @mainthread
    def on_single_conversion_result(self, file_index, tables, sheet_names, error):
        """Callback di main thread setelah satu file selesai diproses."""
        if not (0 <= file_index < len(self.file_list)):
            print(f"Error: Index hasil {file_index} di luar jangkauan saat callback.")
            # Coba proses file berikutnya jika masih ada
            self.current_processing_index = file_index + 1
            self._process_next_file()
            return

        file_info = self.file_list[file_index]
        print(f"Callback hasil untuk file index {file_index}: {file_info['name']}. Error: {error is not None}")

        grid = self.ids.file_list_grid
        current_widget = file_info.get('progress_widget')

        # Cari widget jika referensi hilang
        widget_list_index = (len(self.file_list) - 1 - file_index) * 3
        if not current_widget or current_widget not in grid.children:
            try:
                current_widget = grid.children[widget_list_index]
                if not isinstance(current_widget, (MDLabel, MDBoxLayout)): # Pastikan tipe widget benar
                     current_widget = None
            except IndexError:
                current_widget = None

        new_widget = None
        widget_visual_index = -1
        if current_widget:
             try:
                 widget_visual_index = grid.children.index(current_widget)
                 grid.remove_widget(current_widget)
             except ValueError: # Widget sudah tidak ada di children
                  print(f"Peringatan: Widget progress untuk index {file_index} tidak ditemukan di grid saat callback.")
                  widget_visual_index = widget_list_index # Gunakan index kalkulasi
        else:
             print(f"Error: Tidak dapat menemukan widget progress sama sekali untuk index {file_index} saat callback.")
             widget_visual_index = widget_list_index # Gunakan index kalkulasi


        if error:
            file_info['status'] = 'failed'
            new_widget = FailedStatusCell(size_hint_x=0.25)
            print(f"File index {file_index} ditandai GAGAL.")
        elif tables and all(isinstance(df, pd.DataFrame) for df in tables):
            # Sukses, ada tabel ditemukan
            file_info['tables'] = tables
            file_info['sheet_names'] = sheet_names
            try:
                # Buat file output di memori sesuai format yang dipilih
                output_bytes, _, file_ext = create_tables_export(tables, self.selected_format, sheet_names)
                if output_bytes:
                    file_info['output_data'] = output_bytes
                    file_info['file_ext'] = file_ext
                    file_info['status'] = 'success'
                    new_widget = DownloadButtonCell(file_index=file_info['index'], size_hint_x=0.25)
                    print(f"File index {file_index} SUKSES diekspor ke {file_ext}.")
                else:
                    raise ValueError(f"Ekspor ke {self.selected_format} menghasilkan data kosong.")
            except Exception as e:
                print(f"Gagal membuat file ekspor untuk index {file_index}: {e}")
                file_info['status'] = 'failed'
                file_info['output_data'] = None
                new_widget = FailedStatusCell(size_hint_x=0.25)
                print(f"File index {file_index} ditandai GAGAL saat ekspor.")
        else:
            # Sukses, tapi tidak ada tabel ditemukan atau format tabel salah
            file_info['status'] = 'no_tables'
            new_widget = NoTablesStatusCell(size_hint_x=0.25)
            print(f"File index {file_index} SUKSES tapi tidak ada tabel.")

        # Tambahkan widget status baru ke grid jika berhasil dibuat & index valid
        if new_widget and widget_visual_index != -1:
             try:
                 grid.add_widget(new_widget, index=widget_visual_index)
                 file_info['progress_widget'] = new_widget # Update referensi
             except Exception as e:
                  print(f"Error menambahkan widget hasil ke grid index {widget_visual_index}: {e}")
                  # Coba tambahkan di akhir jika index bermasalah
                  try: grid.add_widget(new_widget)
                  except: print("Gagal menambahkan widget hasil sama sekali.")


        # Lanjutkan ke file berikutnya
        self.current_processing_index = file_index + 1
        self._process_next_file()


    def _prepare_and_go_to_result_screen(self):
         """Menyiapkan UI layar hasil dan beralih ke sana."""
         result_screen = self.manager.result_screen
         if not result_screen:
             print("Error: Layar hasil tidak ditemukan.")
             self.reset_ui_state() # Kembali ke awal jika layar hasil hilang
             return

         result_grid = result_screen.ids.result_file_list_grid
         result_grid.clear_widgets() # Kosongkan hasil sebelumnya

         any_success = False
         success_count = 0

         for f_info in self.file_list:
             any_success = any_success or (f_info['status'] == 'success')
             if f_info['status'] == 'success': success_count += 1

             # Tambahkan baris ke grid hasil
             result_grid.add_widget(ResultCellLabel(text=f_info['name'], size_hint_x=0.6))
             result_grid.add_widget(ResultCellLabelRight(text=f_info['size'], size_hint_x=0.15))

             status_widget = None
             if f_info['status'] == 'success':
                 status_widget = ResultSuccessCell(file_index=f_info['index'], size_hint_x=0.25)
             elif f_info['status'] == 'failed':
                 status_widget = ResultFailedCell(size_hint_x=0.25)
             else: # 'no_tables' or 'pending' (seharusnya tidak pending)
                 status_widget = ResultNoTablesCell(size_hint_x=0.25)
             result_grid.add_widget(status_widget)

         # Update label status dan tombol Unduh Semua di layar hasil
         result_screen.ids.save_button_result.disabled = not any_success
         if success_count > 1:
             result_screen.ids.save_button_result.text = 'Unduh Semua (ZIP)'
         else:
             result_screen.ids.save_button_result.text = 'Unduh Semua'

         if any_success:
             result_screen.ids.result_status_label.text = f"{success_count} dari {len(self.file_list)} file berhasil. Silakan unduh."
         elif any(f['status'] == 'no_tables' for f in self.file_list):
              result_screen.ids.result_status_label.text = "Proses selesai, namun tidak ada tabel ditemukan di file manapun."
         else: # Semua gagal
             result_screen.ids.result_status_label.text = "Semua file gagal diproses."

         # Pindah ke layar hasil
         self.manager.go_to_result_screen()


    def download_single_file(self, file_index):
        """Memulai proses unduh untuk satu file yang berhasil."""
        if not (0 <= file_index < len(self.file_list)):
            self.show_alert("Error", "Index file tidak valid.")
            return

        file_info = self.file_list[file_index]

        if file_info['status'] != 'success' or not file_info['output_data']:
            self.show_alert("Info", f"File '{file_info['name']}' belum siap atau gagal dikonversi.")
            return

        output_data = file_info['output_data']
        file_ext = file_info.get('file_ext', self.selected_format or 'xlsx') # Fallback extension

        # Jika formatnya CSV dan ada >1 tabel di file INI, zip otomatis
        is_multitable_csv = (file_ext == 'csv' and
                             file_info['tables'] and
                             isinstance(file_info['tables'], list) and
                             len(file_info['tables']) > 1)

        temp_output_data = output_data
        temp_file_ext = file_ext

        if is_multitable_csv:
             print(f"File CSV '{file_info['name']}' memiliki {len(file_info['tables'])} tabel, akan di-zip.")
             zip_buffer = io.BytesIO()
             try:
                 with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                      sheet_names = file_info.get('sheet_names', [])
                      for i, df in enumerate(file_info['tables']):
                           # Buat nama file di dalam zip
                           sheet_name = sheet_names[i] if i < len(sheet_names) else f"Tabel_{i+1}"
                           safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
                           filename_in_zip = f"{safe_sheet_name}.csv"
                           zf.writestr(filename_in_zip, df.to_csv(index=False, header=False, encoding='utf-8-sig'))
                 temp_output_data = zip_buffer.getvalue()
                 temp_file_ext = 'zip' # Ubah ekstensi menjadi zip
             except Exception as e:
                  print(f"Gagal membuat zip untuk multi-tabel CSV: {e}")
                  self.show_alert("Error", "Gagal membuat file ZIP untuk hasil CSV.")
                  return


        # Buat nama file default dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"{file_info['original_name']}_{timestamp}.{temp_file_ext}"

        # Buka dialog simpan file
        try:
            filechooser.save_file(
                on_selection=lambda path: self.write_saved_file(path, temp_output_data, expected_ext=temp_file_ext),
                path=default_filename # Sugesti nama file
            )
        except Exception as e:
            print(f"Gagal membuka dialog simpan file: {e}")
            self.show_alert("Error", "Gagal membuka dialog simpan file.")

    def save_result(self):
        """Menyimpan semua file yang berhasil, di-zip jika lebih dari satu."""
        successful_files = [f for f in self.file_list if f['status'] == 'success' and f.get('output_data')]

        if not successful_files:
            self.show_alert("Info", "Tidak ada file yang berhasil dikonversi untuk disimpan.")
            return

        final_output_data = None
        final_file_ext = self.selected_format or 'xlsx' # Default jika belum dipilih
        base_filename = "hasil_konversi"

        if len(successful_files) == 1:
            # Unduh sebagai file tunggal (panggil fungsi download_single_file)
            self.download_single_file(successful_files[0]['index'])
            return

        # Jika lebih dari 1 file sukses, buat ZIP
        final_file_ext = 'zip'
        base_filename = "hasil_konversi_batch"
        zip_buffer = io.BytesIO()
        print("Membuat file ZIP untuk semua hasil...")

        try:
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file_info in successful_files:
                    output_data_single = file_info['output_data']
                    file_ext_single = file_info.get('file_ext', self.selected_format or 'xlsx')
                    timestamp_individual = datetime.now().strftime("%H%M%S") # Timestamp unik per file

                    is_multitable_csv = (file_ext_single == 'csv' and
                                         file_info['tables'] and
                                         isinstance(file_info['tables'], list) and
                                         len(file_info['tables']) > 1)

                    if is_multitable_csv:
                        # Jika CSV multi-tabel, pecah di dalam zip
                         sheet_names = file_info.get('sheet_names', [])
                         for i, df in enumerate(file_info['tables']):
                              sheet_name = sheet_names[i] if i < len(sheet_names) else f"Tabel_{i+1}"
                              safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
                              # Nama file: original_timestamp_sheetname.csv
                              filename_in_zip = f"{file_info['original_name']}_{timestamp_individual}_{safe_sheet_name}.csv"
                              try:
                                   zf.writestr(filename_in_zip, df.to_csv(index=False, header=False, encoding='utf-8-sig'))
                              except Exception as e_csv:
                                   print(f"Gagal menulis sheet {filename_in_zip} ke zip: {e_csv}")
                    elif output_data_single:
                        # Nama file: original_timestamp.ext
                        filename_in_zip = f"{file_info['original_name']}_{timestamp_individual}.{file_ext_single}"
                        zf.writestr(filename_in_zip, output_data_single)
                    else:
                         print(f"Skipping file {file_info['name']} karena tidak ada data output.")

            final_output_data = zip_buffer.getvalue()
        except Exception as e:
            print(f"Gagal membuat file ZIP: {e}")
            self.show_alert("Error", f"Gagal membuat file ZIP: {e}")
            return

        if not final_output_data:
            self.show_alert("Error", "Gagal membuat file ZIP (data kosong).")
            return

        # Buka dialog simpan file untuk ZIP
        timestamp_outer = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"{base_filename}_{timestamp_outer}.{final_file_ext}"

        try:
            filechooser.save_file(
                on_selection=lambda path: self.write_saved_file(path, final_output_data, expected_ext=final_file_ext),
                path=default_filename
            )
        except Exception as e:
            print(f"Gagal membuka dialog simpan file ZIP: {e}")
            self.show_alert("Error", "Gagal membuka dialog simpan file.")


    def write_saved_file(self, path_selection, data, expected_ext):
        """Menulis data bytes ke file yang dipilih pengguna."""
        if not path_selection: return # Pengguna membatalkan

        try:
            # Hasil filechooser bisa berupa list (jika multiple=True, tapi di save=False harusnya string)
            save_path = path_selection[0] if isinstance(path_selection, (list, tuple)) else path_selection

            # Pastikan ekstensi file benar
            base, ext = os.path.splitext(save_path)
            clean_expected_ext = expected_ext.lstrip('.').lower()
            current_ext = ext.lstrip('.').lower()

            if current_ext != clean_expected_ext:
                save_path = f"{base}.{clean_expected_ext}"
                print(f"Ekstensi file diperbaiki menjadi: {save_path}")

            with open(save_path, 'wb') as f:
                f.write(data)

            status_msg = f"File berhasil disimpan: {os.path.basename(save_path)}"
            self.show_alert("Sukses", status_msg)
            print(status_msg)
            # Update status di layar aktif
            if self.manager.current == 'main':
                 self.ids.status_label.text = status_msg
            elif self.manager.current == 'result':
                 self.manager.result_screen.ids.result_status_label.text = status_msg

        except Exception as e:
            status_msg = f"Gagal menyimpan file: {e}"
            self.show_alert("Error Penyimpanan", status_msg)
            print(status_msg)
            if self.manager.current == 'main':
                 self.ids.status_label.text = status_msg
            elif self.manager.current == 'result':
                 self.manager.result_screen.ids.result_status_label.text = status_msg
            import traceback
            traceback.print_exc()


# --- Definisi Kelas Screen Manager ---
class RootScreenManager(MDScreenManager):
    """Pengelola layar utama, mengatur transisi antar layar."""
    main_screen = ObjectProperty(None)
    result_screen = ObjectProperty(None)

    def go_to_main_screen(self):
        """Kembali ke layar utama dan mereset statusnya."""
        if self.main_screen:
            self.main_screen.reset_ui_state()
        if self.result_screen:
             # Kosongkan grid hasil saat kembali
             try: # Tambahkan try-except jika ids belum tentu ada
                  self.result_screen.ids.result_file_list_grid.clear_widgets()
             except AttributeError:
                  pass
        self.current = 'main'
        print("Navigasi ke Layar Utama.")

    def go_to_result_screen(self):
        """Pindah ke layar hasil."""
        self.current = 'result'
        print("Navigasi ke Layar Hasil.")


# --- Definisi Kelas Aplikasi Utama ---
class PDFExtractApp(MDApp):
    """Kelas aplikasi KivyMD utama."""
    def build(self):
        """Membangun UI aplikasi."""
        self.title = "PDFExtract" # Judul jendela aplikasi

        # Coba atur ikon aplikasi
        try:
            icon_path = os.path.join(application_path, 'icon.ico') # Ganti nama jika perlu
            if os.path.exists(icon_path):
                self.icon = icon_path
            else:
                # Coba cari .png jika .ico tidak ada
                icon_path_png = os.path.join(application_path, 'icon.png')
                if os.path.exists(icon_path_png):
                     self.icon = icon_path_png
                else:
                     print(f"Peringatan: File ikon tidak ditemukan di {icon_path} atau {icon_path_png}")
        except Exception as e:
            print(f"Gagal mengatur ikon aplikasi: {e}")

        # Atur tema warna
        self.theme_cls.primary_palette = "Blue"  # Warna utama (misal, AppBar)
        self.theme_cls.accent_palette = "Green" # Warna aksen (misal, tombol penting)
        self.theme_cls.theme_style = "Light"     # Tema terang

        # Atur font default (opsional)
        # self.theme_cls.font_styles["Regular"] = "Roboto" # Ganti dengan path font jika perlu

        return RootScreenManager() # Kembalikan instance Screen Manager sebagai root widget

    def on_stop(self):
        """Dipanggil saat aplikasi ditutup."""
        print("Aplikasi ditutup.")
        # Tambahkan pembersihan sumber daya jika perlu di sini


# --- Entry Point Aplikasi ---
if __name__ == '__main__':
    # Jalankan aplikasi
    PDFExtractApp().run()